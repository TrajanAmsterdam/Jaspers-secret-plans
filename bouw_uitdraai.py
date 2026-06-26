# -*- coding: utf-8 -*-
"""
Bouwt een formule-gedreven werkbestand voor de stromen-/intensiteitenuitdraai.
Leest de ruwe uurdata uit het rapportagebestand (read-only kopie) en schrijft een
NIEUW bestand met live Excel-formules (SUMIFS/AVERAGEIFS), zodat alles controleerbaar is.
"""
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

SRC = os.path.join(os.environ["TEMP"], "_stromen_kopie.xlsx")
OUT = r"D:\Trajan B.V\1620.26.1 Capelle aan de IJssel Amerikaansebuurt - Intern\Rapportage\1620.26.1 intensiteiten-uitdraai (formules).xlsx"

WERKDAGEN = {"Maandag", "Dinsdag", "Woensdag", "Donderdag", "Vrijdag"}

# Richting-mapping: template-kolomlabel -> richtingstring in de data
# "Richting X" = verkeer NAAR X toe (eindigt op 'richting X')
RICHTING_LABELS = {
    "Noorderbreedte": [
        ("Richting Hellingweg", "Purmerhoek richting Hellingweg"),
        ("Richting Purmerhoek", "Hellingweg richting Purmerhoek"),
    ],
    "Westerlengte": [
        ("Richting Floridaweg", "Bisletweg richting Floridaweg"),
        ("Richting Bisletweg", "Floridaweg richting Bisletweg"),
    ],
}

TIJDVAKKEN = [  # (label, agg-kolomnaam)
    ("Etmaal (0-24u)", "Etmaal"),
    ("Dag (7-19u)", "Dag"),
    ("Avond (19-23u)", "Avond"),
    ("Nacht (23-7u)", "Nacht"),
    ("Ochtendspits (7-9u)", "OchtSpits"),
    ("Avondspits (16-18u)", "AvondSpits"),
]

# ---------------------------------------------------------------- inlezen
src = openpyxl.load_workbook(SRC, read_only=True, data_only=True)


def lees(sheet, nkol):
    ws = src[sheet]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None or r[4] is None:
            continue
        rows.append(list(r[:nkol]))
    return rows


moto = lees("moto-samengevoegd", 32)   # A..AF (32 kolommen)
fiets = lees("fiets-samengevoegd", 6)   # A..F  (6 kolommen)
src.close()
print(f"moto rijen: {len(moto)} | fiets rijen: {len(fiets)}")


def uur(tijd):
    return int(str(tijd)[:2])


# ---------------------------------------------------------------- nieuw wb
wb = Workbook()
wb.remove(wb.active)

bold = Font(bold=True)
hdr_fill = PatternFill("solid", fgColor="DDEBF7")
sub_fill = PatternFill("solid", fgColor="F2F2F2")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center")

# ============================ data-moto ============================
wsm = wb.create_sheet("data-moto")
moto_hdr = ["Locatie", "Datum", "Dag", "Tijd", "Richting",
            "<3,7;<30", "<3,7;30-40", "<3,7;40-50", "<3,7;50-55", "<3,7;55-60",
            "<3,7;60-70", "<3,7;70-80", "<3,7;>=80",
            "3,7-7;<30", "3,7-7;30-40", "3,7-7;40-50", "3,7-7;50-55", "3,7-7;55-60",
            "3,7-7;60-70", "3,7-7;70-80", "3,7-7;>=80",
            ">=7;<30", ">=7;30-40", ">=7;40-50", ">=7;50-55", ">=7;55-60",
            ">=7;60-70", ">=7;70-80", ">=7;>=80",
            "V85", "Gem. snelheid", "Totaal",
            "Uur", "Licht", "Middel", "Zwaar"]
wsm.append(moto_hdr)
for r in moto:
    wsm.append(r)
# hulpkolommen (AG=33 Uur, AH=34 Licht, AI=35 Middel, AJ=36 Zwaar)
for i, r in enumerate(moto):
    row = i + 2
    wsm.cell(row=row, column=33, value=uur(r[3]))                       # Uur
    wsm.cell(row=row, column=34, value=f"=SUM(F{row}:M{row})")          # Licht  <3,7
    wsm.cell(row=row, column=35, value=f"=SUM(N{row}:U{row})")          # Middel 3,7-7
    wsm.cell(row=row, column=36, value=f"=SUM(V{row}:AC{row})")         # Zwaar  >=7
moto_last = len(moto) + 1
for c in range(1, 37):
    wsm.cell(row=1, column=c).font = bold

# ============================ data-fiets ===========================
wsf = wb.create_sheet("data-fiets")
wsf.append(["Locatie", "Datum", "Dag", "Tijd", "Richting", "Aantal", "Uur"])
for r in fiets:
    wsf.append(r)
for i, r in enumerate(fiets):
    row = i + 2
    wsf.cell(row=row, column=7, value=uur(r[3]))   # Uur (kolom G)
fiets_last = len(fiets) + 1
for c in range(1, 8):
    wsf.cell(row=1, column=c).font = bold

# ============================ agg ==================================
wsa = wb.create_sheet("agg")
agg_hdr = ["Modaliteit", "Locatie", "Richting", "Datum", "Dag", "Werkdag",
           "Etmaal", "Dag", "Avond", "Nacht", "OchtSpits", "AvondSpits",
           "Licht", "Middel", "Zwaar"]
wsa.append(agg_hdr)
for c in range(1, 16):
    wsa.cell(row=1, column=c).font = bold

# unieke combinaties (modaliteit, loc, richting, datum, dag) in databron-volgorde
def combos(rows):
    seen = {}
    for r in rows:
        key = (r[0], r[4], r[1])  # loc, richting, datum
        if key not in seen:
            seen[key] = r[2]       # dag
    return [(loc, ri, dt, dag) for (loc, ri, dt), dag in seen.items()]

agg_row = 2
for mod, rows, last, dsheet, valcol in [
    ("moto", moto, moto_last, "data-moto", "AF"),
    ("fiets", fiets, fiets_last, "data-fiets", "F"),
]:
    # kolomletters in de databron
    if mod == "moto":
        c_loc, c_ri, c_dat, c_uur = "A", "E", "B", "AG"
        c_licht, c_mid, c_zw = "AH", "AI", "AJ"
    else:
        c_loc, c_ri, c_dat, c_uur = "A", "E", "B", "G"
    rng = lambda col: f"'{dsheet}'!${col}$2:${col}${last}"
    val = rng(valcol)
    for (loc, ri, dt, dag) in combos(rows):
        R = agg_row
        wsa.cell(row=R, column=1, value=mod)
        wsa.cell(row=R, column=2, value=loc)
        wsa.cell(row=R, column=3, value=ri)
        c = wsa.cell(row=R, column=4, value=dt); c.number_format = "DD-MM-YYYY"
        wsa.cell(row=R, column=5, value=dag)
        wsa.cell(row=R, column=6, value=1 if dag in WERKDAGEN else 0)
        # criteria die overal gelden:
        crit = f"{rng(c_loc)},$B{R},{rng(c_ri)},$C{R},{rng(c_dat)},$D{R}"
        # Etmaal (alle uren)
        wsa.cell(row=R, column=7, value=f"=SUMIFS({val},{crit})")
        # Dag 7-18
        wsa.cell(row=R, column=8, value=f"=SUMIFS({val},{crit},{rng(c_uur)},\">=7\",{rng(c_uur)},\"<=18\")")
        # Avond 19-22
        wsa.cell(row=R, column=9, value=f"=SUMIFS({val},{crit},{rng(c_uur)},\">=19\",{rng(c_uur)},\"<=22\")")
        # Nacht = Etmaal - Dag - Avond
        wsa.cell(row=R, column=10, value=f"=G{R}-H{R}-I{R}")
        # Ochtendspits 7-8
        wsa.cell(row=R, column=11, value=f"=SUMIFS({val},{crit},{rng(c_uur)},\">=7\",{rng(c_uur)},\"<=8\")")
        # Avondspits 16-17
        wsa.cell(row=R, column=12, value=f"=SUMIFS({val},{crit},{rng(c_uur)},\">=16\",{rng(c_uur)},\"<=17\")")
        if mod == "moto":
            wsa.cell(row=R, column=13, value=f"=SUMIFS({rng(c_licht)},{crit})")
            wsa.cell(row=R, column=14, value=f"=SUMIFS({rng(c_mid)},{crit})")
            wsa.cell(row=R, column=15, value=f"=SUMIFS({rng(c_zw)},{crit})")
        agg_row += 1
agg_last = agg_row - 1
print(f"agg rijen: {agg_last - 1}")

# kolomnaam -> agg kolomletter
AGG = {"Etmaal": "G", "Dag": "H", "Avond": "I", "Nacht": "J",
       "OchtSpits": "K", "AvondSpits": "L", "Licht": "M", "Middel": "N", "Zwaar": "O"}


def aref(col):
    return f"agg!${col}$2:${col}${agg_last}"


# ============================ Uitdraai tabellen ====================
ws = wb.create_sheet("Uitdraai tabellen")
ws.sheet_view.showGridLines = False
for col, w in {"A": 20, "B": 11, "C": 9, "D": 11, "E": 9, "F": 12, "G": 12, "H": 12, "I": 12}.items():
    ws.column_dimensions[col].width = w


def schrijf_intensiteit(start, titel, mod, loc):
    """Schrijft een intensiteiten-tabel (6 tijdvakken) vanaf rij `start`. Geeft eindrij terug."""
    labels = RICHTING_LABELS[loc]
    r1lab, r1str = labels[0]
    r2lab, r2str = labels[1]
    base = f'agg!$A$2:$A${agg_last},"{mod}",agg!$B$2:$B${agg_last},"{loc}",agg!$C$2:$C${agg_last},'

    # titelrij
    ws.cell(row=start, column=1, value=titel).font = Font(bold=True, size=12)
    ws.cell(row=start, column=2, value=loc).font = bold
    # kopregel 1
    h = start + 1
    ws.cell(row=h, column=1, value="Intensiteiten").font = bold
    ws.cell(row=h, column=2, value="Doorsnede").font = bold
    ws.cell(row=h, column=4, value="").font = bold
    ws.cell(row=h, column=6, value=r1lab).font = bold
    ws.cell(row=h, column=8, value=r2lab).font = bold
    ws.merge_cells(start_row=h, start_column=2, end_row=h, end_column=5)
    ws.merge_cells(start_row=h, start_column=6, end_row=h, end_column=7)
    ws.merge_cells(start_row=h, start_column=8, end_row=h, end_column=9)
    # kopregel 2
    s = start + 2
    sub = ["Tijdvak", "Werkdag", "", "Weekdag", "", "Werkdag", "Weekdag", "Werkdag", "Weekdag"]
    for j, t in enumerate(sub):
        c = ws.cell(row=s, column=1 + j, value=t)
        c.font = bold; c.alignment = center; c.fill = sub_fill
    ws.merge_cells(start_row=s, start_column=2, end_row=s, end_column=3)
    ws.merge_cells(start_row=s, start_column=4, end_row=s, end_column=5)
    # kopregel 3 (Aantal/aandeel)
    s2 = start + 3
    sub2 = ["", "Aantal", "aandeel", "Aantal", "aandeel", "Aantal", "Aantal", "Aantal", "Aantal"]
    for j, t in enumerate(sub2):
        c = ws.cell(row=s2, column=1 + j, value=t)
        c.font = bold; c.alignment = center; c.fill = sub_fill
    # datarijen
    first_data = start + 4
    etmaal_row = first_data  # Etmaal is de eerste tijdvakrij (voor aandeel)
    for k, (tlab, tcol) in enumerate(TIJDVAKKEN):
        rr = first_data + k
        ac = aref(AGG[tcol])
        ws.cell(row=rr, column=1, value=tlab)
        # richtingen (AVERAGEIFS)
        f = f'=AVERAGEIFS({ac},{base}"{r1str}",agg!$F$2:$F${agg_last},1)'      # R1 werkdag
        g = f'=AVERAGEIFS({ac},{base}"{r1str}")'                                # R1 weekdag
        hh = f'=AVERAGEIFS({ac},{base}"{r2str}",agg!$F$2:$F${agg_last},1)'      # R2 werkdag
        ii = f'=AVERAGEIFS({ac},{base}"{r2str}")'                               # R2 weekdag
        ws.cell(row=rr, column=6, value=f)
        ws.cell(row=rr, column=7, value=g)
        ws.cell(row=rr, column=8, value=hh)
        ws.cell(row=rr, column=9, value=ii)
        # doorsnede = R1 + R2
        ws.cell(row=rr, column=2, value=f"=F{rr}+H{rr}")   # werkdag aantal
        ws.cell(row=rr, column=4, value=f"=G{rr}+I{rr}")   # weekdag aantal
        # aandeel t.o.v. etmaal
        ws.cell(row=rr, column=3, value=f"=IF(B${etmaal_row}=0,\"\",B{rr}/B${etmaal_row})").number_format = "0.0%"
        ws.cell(row=rr, column=5, value=f"=IF(D${etmaal_row}=0,\"\",D{rr}/D${etmaal_row})").number_format = "0.0%"
        for col in (2, 4, 6, 7, 8, 9):
            ws.cell(row=rr, column=col).number_format = "#,##0"
    return first_data + len(TIJDVAKKEN) - 1


def schrijf_voertuig(start, titel, loc):
    """Voertuigverdeling (L/M/Z), etmaal-niveau. Geeft eindrij terug."""
    labels = RICHTING_LABELS[loc]
    r1lab, r1str = labels[0]
    r2lab, r2str = labels[1]
    base = f'agg!$A$2:$A${agg_last},"moto",agg!$B$2:$B${agg_last},"{loc}",agg!$C$2:$C${agg_last},'

    ws.cell(row=start, column=1, value=titel).font = Font(bold=True, size=12)
    ws.cell(row=start, column=2, value=loc).font = bold
    h = start + 1
    ws.cell(row=h, column=1, value="Voertuigverdeling").font = bold
    ws.cell(row=h, column=2, value="Doorsnede").font = bold
    ws.cell(row=h, column=6, value=r1lab).font = bold
    ws.cell(row=h, column=8, value=r2lab).font = bold
    ws.merge_cells(start_row=h, start_column=2, end_row=h, end_column=5)
    ws.merge_cells(start_row=h, start_column=6, end_row=h, end_column=7)
    ws.merge_cells(start_row=h, start_column=8, end_row=h, end_column=9)
    s = start + 2
    sub = ["Modaliteit", "Werkdag", "", "Weekdag", "", "Werkdag", "Weekdag", "Werkdag", "Weekdag"]
    for j, t in enumerate(sub):
        c = ws.cell(row=s, column=1 + j, value=t); c.font = bold; c.alignment = center; c.fill = sub_fill
    ws.merge_cells(start_row=s, start_column=2, end_row=s, end_column=3)
    ws.merge_cells(start_row=s, start_column=4, end_row=s, end_column=5)
    s2 = start + 3
    sub2 = ["", "Aantal", "aandeel", "Aantal", "aandeel", "Aantal", "Aantal", "Aantal", "Aantal"]
    for j, t in enumerate(sub2):
        c = ws.cell(row=s2, column=1 + j, value=t); c.font = bold; c.alignment = center; c.fill = sub_fill
    first_data = start + 4
    mods = [("Licht (L)", "Licht"), ("Middel (M)", "Middel"), ("Zwaar (Z)", "Zwaar")]
    rows_idx = list(range(first_data, first_data + 3))
    for k, (mlab, mcol) in enumerate(mods):
        rr = first_data + k
        ac = aref(AGG[mcol])
        ws.cell(row=rr, column=1, value=mlab)
        ws.cell(row=rr, column=6, value=f'=AVERAGEIFS({ac},{base}"{r1str}",agg!$F$2:$F${agg_last},1)')
        ws.cell(row=rr, column=7, value=f'=AVERAGEIFS({ac},{base}"{r1str}")')
        ws.cell(row=rr, column=8, value=f'=AVERAGEIFS({ac},{base}"{r2str}",agg!$F$2:$F${agg_last},1)')
        ws.cell(row=rr, column=9, value=f'=AVERAGEIFS({ac},{base}"{r2str}")')
        ws.cell(row=rr, column=2, value=f"=F{rr}+H{rr}")
        ws.cell(row=rr, column=4, value=f"=G{rr}+I{rr}")
        for col in (2, 4, 6, 7, 8, 9):
            ws.cell(row=rr, column=col).number_format = "#,##0"
    # aandeel = aandeel binnen L+M+Z per kolom
    f, l = rows_idx[0], rows_idx[-1]
    for rr in rows_idx:
        ws.cell(row=rr, column=3, value=f"=IF(SUM(B${f}:B${l})=0,\"\",B{rr}/SUM(B${f}:B${l}))").number_format = "0.0%"
        ws.cell(row=rr, column=5, value=f"=IF(SUM(D${f}:D${l})=0,\"\",D{rr}/SUM(D${f}:D${l}))").number_format = "0.0%"
    return first_data + 2


# --- plaatsing van de blokken ---
r = 1
r = schrijf_intensiteit(r, "FIETS", "fiets", "Noorderbreedte") + 3
r = schrijf_intensiteit(r, "FIETS", "fiets", "Westerlengte") + 3
r = schrijf_intensiteit(r, "AUTO", "moto", "Noorderbreedte") + 3
r = schrijf_intensiteit(r, "AUTO", "moto", "Westerlengte") + 3
r = schrijf_voertuig(r, "AUTO voertuigverdeling", "Noorderbreedte") + 3
r = schrijf_voertuig(r, "AUTO voertuigverdeling", "Westerlengte") + 3

# ============================ Toelichting ==========================
wt = wb.create_sheet("Toelichting")
wt.column_dimensions["A"].width = 28
wt.column_dimensions["B"].width = 90
toel = [
    ("Onderwerp", "Uitleg"),
    ("Doel", "Geaggregeerde intensiteiten per tijdvak, dagtype en richting voor fiets en gemotoriseerd verkeer."),
    ("Bron", "Tabbladen 'data-moto' en 'data-fiets' = kopie van de ruwe uurdata uit het rapportagebestand."),
    ("Tijdvakgrenzen", "Op beginuur. Dag=07-18u, Avond=19-22u, Ochtendspits=07-08u, Avondspits=16-17u, Etmaal=00-23u."),
    ("Nacht (23-7u)", "Berekend als Etmaal - Dag - Avond (= uren 23 t/m 06). Zo telt de nacht over middernacht correct mee."),
    ("Werkdag", "Gemiddelde over maandag t/m vrijdag (kolom 'Dag')."),
    ("Weekdag", "Gemiddelde over alle 7 dagen (ma t/m zo)."),
    ("Aantal", "Gemiddeld dagtotaal binnen het tijdvak = AVERAGE over de dagen van het dagtotaal (SUMIFS) per dag (zie tab 'agg')."),
    ("Doorsnede", "Som van beide richtingen (= richting 1 + richting 2)."),
    ("aandeel (intens.)", "Aandeel van het tijdvak t.o.v. het etmaal (0-24u) binnen dezelfde kolom."),
    ("Licht / Middel / Zwaar", "Licht=lengte <3,7m (som F:M), Middel=3,7-7m (som N:U), Zwaar=>=7m (som V:AC) in 'data-moto'."),
    ("aandeel (voertuig)", "Aandeel van L/M/Z binnen het etmaaltotaal L+M+Z per kolom."),
    ("Geldige dagen", "Alleen datums waarvoor data bestaat (Noorderbreedte 17 dagen, Westerlengte 16 dagen). Alle dagen 24u compleet."),
    ("Richting-mapping", "'Richting X' = verkeer naar X toe. Pas zonodig RICHTING_LABELS aan in het bouwscript."),
]
for i, (a, b) in enumerate(toel):
    wt.cell(row=i + 1, column=1, value=a)
    wt.cell(row=i + 1, column=2, value=b)
    if i == 0:
        wt.cell(row=1, column=1).font = bold
        wt.cell(row=1, column=2).font = bold
    wt.cell(row=i + 1, column=2).alignment = Alignment(wrap_text=True, vertical="top")

# volgorde tabbladen: Uitdraai eerst
wb.move_sheet("Uitdraai tabellen", -(wb.sheetnames.index("Uitdraai tabellen")))
wb.move_sheet("Toelichting", 1 - wb.sheetnames.index("Toelichting"))

wb.save(OUT)
print("OPGESLAGEN:", OUT)
