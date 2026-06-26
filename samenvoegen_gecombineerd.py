import sys
import csv
import re
from pathlib import Path
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

# Optionele onderdelen: slepen-en-neerzetten en kaart. Ontbreken ze, dan
# blijft de tool werken (zonder die functie).
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    HAS_DND = True
except Exception:
    HAS_DND = False

try:
    from tkintermapview import TkinterMapView
    HAS_MAP = True
except Exception:
    HAS_MAP = False

try:
    import win32com.client as win32
    HAS_COM = True
except Exception:
    HAS_COM = False


# Voertuigtype-labels in de uitvoer.
FIETS_VOERTUIGTYPE = "fiets"
TELSLANG_VOERTUIGTYPE = "motorvoertuig"

# Standaard dagdeel-indeling (van inclusief, tot exclusief; van>tot loopt over middernacht)
DEFAULT_DAGDELEN = [
    ("07:00", "10:00", "ochtendspits"),
    ("10:00", "16:00", "overdag"),
    ("16:00", "19:00", "avondspits"),
    ("19:00", "07:00", "nacht"),
]

DAG_KOLOMMEN = {
    "Maandag": 0, "Dinsdag": 1, "Woensdag": 2, "Donderdag": 3,
    "Vrijdag": 4, "Zaterdag": 5, "Zondag": 6,
}

WEEK_HEADER = re.compile(r"^(\d{2}-\d{2}-\d{4})\s*-\s*(\d{2}-\d{2}-\d{4})\s*-\s*(.+)$")
DAG_HEADER = re.compile(r"^(\d{2}-\d{2}-\d{4})\s*-\s*(.+)$")


# ---------------------------------------------------------------------------
# Dagdelen / dagsoort
# ---------------------------------------------------------------------------

def basis_dir():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent


def _tijd_naar_minuten(waarde):
    m = re.match(r"^\s*(\d{1,2}):(\d{2})\s*$", str(waarde))
    if not m:
        return None
    return int(m.group(1)) * 60 + int(m.group(2))


def lees_dagdelen_strings(base_dir):
    pad = base_dir / "dagdelen.csv"
    rijen = []
    try:
        with open(pad, "r", encoding="utf-8-sig") as f:
            for row in csv.reader(f, delimiter=";"):
                if not row or not str(row[0]).strip() or str(row[0]).strip().startswith("#"):
                    continue
                if len(row) < 3:
                    continue
                van, tot, naam = str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip()
                if _tijd_naar_minuten(van) is None or _tijd_naar_minuten(tot) is None or not naam:
                    continue
                rijen.append((van, tot, naam))
    except OSError:
        pass
    return rijen if rijen else [tuple(r) for r in DEFAULT_DAGDELEN]


def schrijf_dagdelen_strings(base_dir, rijen):
    try:
        with open(base_dir / "dagdelen.csv", "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(["van", "tot", "dagdeel"])
            for van, tot, naam in rijen:
                w.writerow([van, tot, naam])
    except OSError:
        pass


def strings_naar_dagdelen(rijen):
    out = []
    for van, tot, naam in rijen:
        vm, tm = _tijd_naar_minuten(van), _tijd_naar_minuten(tot)
        if vm is not None and tm is not None and naam:
            out.append((vm, tm, naam))
    return out


def load_dagdelen(base_dir):
    return strings_naar_dagdelen(lees_dagdelen_strings(base_dir))


def bepaal_dagdeel(tijd_str, dagdelen):
    t = _tijd_naar_minuten(tijd_str)
    if t is None:
        return ""
    for van, tot, naam in dagdelen:
        if van < tot:
            if van <= t < tot:
                return naam
        elif van > tot:
            if t >= van or t < tot:
                return naam
        else:
            return naam
    return ""


DAGNAMEN = ["maandag", "dinsdag", "woensdag", "donderdag", "vrijdag", "zaterdag", "zondag"]


def dagnaam_van(datum):
    return DAGNAMEN[datum.weekday()]


def bepaal_dagsoort(datum):
    return "weekenddag" if datum.weekday() >= 5 else "weekdag"


def eindtijd_van(tijd_str):
    """Geeft de eindtijd: starttijd + 1 uur. 23:00 -> 24:00."""
    m = _tijd_naar_minuten(tijd_str)
    if m is None:
        return ""
    h, mi = divmod(m + 60, 60)
    return f"{h:02d}:{mi:02d}"


# ---------------------------------------------------------------------------
# Inlezen / parsen
# ---------------------------------------------------------------------------

def read_csv_rows(file_path):
    with open(file_path, "r", encoding="utf-8-sig") as f:
        return list(csv.reader(f, delimiter=";"))


def read_xlsx_rows(file_path, sheet_name):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = [[c if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def to_number(value, empty=""):
    if isinstance(value, (int, float)):
        return value
    if not isinstance(value, str):
        return value
    value = value.strip()
    if not value:
        return empty
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value.replace(",", "."))
    except ValueError:
        return value


def _num(value):
    return value if isinstance(value, (int, float)) else 0


def _cell0(row):
    if row and len(row) > 0 and row[0] is not None:
        return str(row[0]).strip()
    return ""


def parse_rows(rows):
    data = []
    telslang_headers = None
    i, n = 0, len(rows)
    while i < n:
        cell0 = _cell0(rows[i])
        if cell0:
            wm = WEEK_HEADER.match(cell0)
            if wm:
                start_date = datetime.strptime(wm.group(1), "%d-%m-%Y")
                richting = wm.group(3).strip()
                col_header = rows[i + 1] if i + 1 < n else []
                dag_indices = {}
                for j, col in enumerate(col_header):
                    name = str(col).strip()
                    if name in DAG_KOLOMMEN:
                        dag_indices[name] = j
                i += 2
                uur_rijen = []
                while i < n and _cell0(rows[i]):
                    uur_rijen.append(rows[i]); i += 1
                for dag_naam, offset in DAG_KOLOMMEN.items():
                    if dag_naam in dag_indices:
                        datum = start_date + timedelta(days=offset)
                        idx = dag_indices[dag_naam]
                        for dr in uur_rijen:
                            tijd = str(dr[0]).strip()
                            aantal = to_number(dr[idx], empty=0) if idx < len(dr) else 0
                            data.append({"Voertuigtype": FIETS_VOERTUIGTYPE, "Datum": datum,
                                         "Tijd": tijd, "Richting": richting, "classes": {},
                                         "V85": "", "Gem.": "", "Totaal": aantal})
                continue
            dm = DAG_HEADER.match(cell0)
            if dm:
                datum = datetime.strptime(dm.group(1), "%d-%m-%Y")
                richting = dm.group(2).strip()
                header1 = rows[i + 1] if i + 1 < n else []
                header2 = rows[i + 2] if i + 2 < n else []
                block_headers = []
                for j in range(1, 25):
                    lc = str(header1[j]).strip() if j < len(header1) else ""
                    sc = str(header2[j]).strip() if j < len(header2) else ""
                    block_headers.append(f"{lc};{sc}")
                if telslang_headers is None:
                    telslang_headers = block_headers
                i += 3
                while i < n and _cell0(rows[i]):
                    dr = rows[i]
                    tijd = str(dr[0]).strip()
                    classes = {}
                    for k, h in enumerate(block_headers):
                        j = k + 1
                        classes[h] = to_number(dr[j]) if j < len(dr) else ""
                    v85 = to_number(dr[27]) if len(dr) > 27 else ""
                    gem = to_number(dr[29]) if len(dr) > 29 else ""
                    totaal = to_number(dr[30]) if len(dr) > 30 else ""
                    data.append({"Voertuigtype": TELSLANG_VOERTUIGTYPE, "Datum": datum,
                                 "Tijd": tijd, "Richting": richting, "classes": classes,
                                 "V85": v85, "Gem.": gem, "Totaal": totaal})
                    i += 1
                continue
        i += 1
    return data, telslang_headers


def lees_info_velden(path):
    """Lees Code, Naam, Plaats en lat/lon uit het Info-blad van een xlsx.
    Geeft altijd een dict terug (ontbrekende velden zijn None)."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    sheet = next((s for s in wb.sheetnames if s.strip().lower() == "info"), wb.sheetnames[0])
    rows = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    wb.close()

    def f(v):
        try:
            return float(str(v).replace(",", "."))
        except (TypeError, ValueError):
            return None

    code = naam = plaats = lat = lon = None
    for row in rows:
        cells = [("" if c is None else str(c)).strip() for c in row]
        if not cells:
            continue
        sleutel = cells[0].lower()
        if sleutel == "code" and code is None:
            code = next((c for c in cells[1:] if c), None)
        elif sleutel == "naam" and naam is None:
            naam = next((c for c in cells[1:] if c), None)
        elif sleutel == "plaats" and plaats is None:
            plaats = next((c for c in cells[1:] if c), None)
        for i, c in enumerate(cells):
            if c.lower() == "locatie" and i + 2 < len(row):
                a, b = f(row[i + 1]), f(row[i + 2])
                if a is not None and b is not None:
                    lat, lon = a, b
    return {"code": code, "naam": naam, "plaats": plaats, "lat": lat, "lon": lon}


def detecteer_type(path):
    """Snelle detectie of een bestand fietsdata (weektabel) of telslang-data
    (dagtabel) bevat. Geeft FIETS_VOERTUIGTYPE, TELSLANG_VOERTUIGTYPE of None.
    Stopt bij de eerste tabelkop, dus snel."""
    p = Path(path)
    if p.suffix.lower() == ".csv":
        try:
            with open(path, "r", encoding="utf-8-sig") as f:
                for i, row in enumerate(csv.reader(f, delimiter=";")):
                    c0 = str(row[0]).strip() if row and row[0] is not None else ""
                    if c0:
                        if WEEK_HEADER.match(c0):
                            return FIETS_VOERTUIGTYPE
                        if DAG_HEADER.match(c0):
                            return TELSLANG_VOERTUIGTYPE
                    if i > 1000:
                        break
        except OSError:
            return None
        return None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None
    try:
        for s in wb.sheetnames:
            if s.strip().lower() == "info":
                continue
            cnt = 0
            for row in wb[s].iter_rows(values_only=True):
                c0 = str(row[0]).strip() if row and row[0] is not None else ""
                if c0:
                    if WEEK_HEADER.match(c0):
                        return FIETS_VOERTUIGTYPE
                    if DAG_HEADER.match(c0):
                        return TELSLANG_VOERTUIGTYPE
                cnt += 1
                if cnt > 300:
                    break
    finally:
        wb.close()
    return None


def verwerk_bestand(path, verwacht_label):
    """Lees een bestand en geef de rijen van het verwachte type terug.
    Code en Naam komen uit het Info-blad; voor xlsx wordt automatisch het
    juiste data-tabblad gekozen."""
    p = Path(path)
    info = lees_info_velden(path) if p.suffix.lower() == ".xlsx" else None
    code = (info or {}).get("code") or ""
    naam = (info or {}).get("naam") or p.stem

    def _label(rows):
        for d in rows:
            d["Code"] = code
            d["Naam"] = naam
        return rows

    if p.suffix.lower() == ".csv":
        data, headers = parse_rows(read_csv_rows(path))
        typed = _label([d for d in data if d["Voertuigtype"] == verwacht_label])
        return typed, (headers if verwacht_label == TELSLANG_VOERTUIGTYPE else None)

    try:
        wb = load_workbook(path, read_only=True)
        sheetnames = wb.sheetnames
        wb.close()
    except Exception:
        return [], None
    for s in sheetnames:
        if s.strip().lower() == "info":
            continue
        data, headers = parse_rows(read_xlsx_rows(path, s))
        typed = [d for d in data if d["Voertuigtype"] == verwacht_label]
        if typed:
            return _label(typed), headers
    return [], None


# ---------------------------------------------------------------------------
# Uitvoer opbouwen
# ---------------------------------------------------------------------------

def groepeer_klassen(telslang_headers):
    lengte_groepen, snelheid_groepen = {}, {}
    for h in telslang_headers:
        parts = h.split(";")
        lengte = parts[0] if parts else h
        snelheid = parts[1] if len(parts) > 1 else ""
        lengte_groepen.setdefault(lengte, []).append(h)
        snelheid_groepen.setdefault(snelheid, []).append(h)
    return lengte_groepen, snelheid_groepen


def build_output(all_rows, telslang_headers, dagdelen):
    lengte_groepen, snelheid_groepen = ({}, {})
    if telslang_headers:
        lengte_groepen, snelheid_groepen = groepeer_klassen(telslang_headers)

    columns = ["Code", "Naam", "Voertuigtype", "Datum", "Dag", "Dagsoort", "Van", "Tot", "Dagdeel", "Richting"]
    if telslang_headers:
        columns += list(telslang_headers)
        columns += [f"Totaal lengte {l}" for l in lengte_groepen]
        columns += [f"Totaal snelheid {s}" for s in snelheid_groepen]
        columns += ["V85", "Gem."]
    columns += ["Totaal"]

    out_rows = []
    for d in all_rows:
        dagsoort = bepaal_dagsoort(d["Datum"])
        dagdeel = bepaal_dagdeel(d["Tijd"], dagdelen)
        row = [d["Code"], d["Naam"], d["Voertuigtype"], d["Datum"], dagnaam_van(d["Datum"]), dagsoort,
               d["Tijd"], eindtijd_van(d["Tijd"]), dagdeel, d["Richting"]]
        if telslang_headers:
            is_tel = bool(d["classes"])
            for h in telslang_headers:
                row.append(d["classes"].get(h, ""))
            for l, hdrs in lengte_groepen.items():
                row.append(sum(_num(d["classes"].get(h, "")) for h in hdrs) if is_tel else "")
            for s, hdrs in snelheid_groepen.items():
                row.append(sum(_num(d["classes"].get(h, "")) for h in hdrs) if is_tel else "")
            row.append(d["V85"]); row.append(d["Gem."])
        row.append(d["Totaal"])
        out_rows.append(row)
    return columns, out_rows


# Lichte kleuren voor kolomtitels, per lengteklasse op volgorde van voorkomen
KOLOM_KLEUREN = ["FCE4D6", "E2EFDA", "DDEBF7", "FFF2CC", "EAD1DC", "FCE4D6"]


def _lengte_van_kolom(naam):
    """Geeft de lengteklasse van een kolomtitel, of None als die er niet is."""
    if naam.startswith("Totaal lengte "):
        return naam[len("Totaal lengte "):].strip()
    if ";" in naam:
        return naam.split(";")[0].strip()
    return None


def kleur_kolomtitels(ws, out_header):
    """Kleur de kolomtitels (rij 1) per lengteklasse met lichte tinten."""
    volgorde = []
    for naam in out_header:
        lk = _lengte_van_kolom(naam)
        if lk and lk not in volgorde:
            volgorde.append(lk)
    kleur_map = {lk: KOLOM_KLEUREN[i % len(KOLOM_KLEUREN)] for i, lk in enumerate(volgorde)}
    for c, naam in enumerate(out_header, start=1):
        lk = _lengte_van_kolom(naam)
        if lk in kleur_map:
            kl = kleur_map[lk]
            ws.cell(row=1, column=c).fill = PatternFill(
                start_color=kl, end_color=kl, fill_type="solid")


# ---------------------------------------------------------------------------
# Draaitabellen / overzichten op extra tabbladen
# ---------------------------------------------------------------------------

def pivot_specs(out_header):
    """Definieer de overzichtstabbladen op basis van beschikbare kolommen."""
    specs = [
        {"sheet": "Locatie x dagdeel", "rows": ["Naam", "Voertuigtype"],
         "col": "Dagdeel", "data": [("Totaal", "sum", "Som van Totaal")]},
        {"sheet": "Locatie x weekdag-weekend", "rows": ["Naam", "Voertuigtype"],
         "col": "Dagsoort", "data": [("Totaal", "sum", "Som van Totaal")]},
        {"sheet": "Etmaal per locatie per dag", "rows": ["Naam", "Datum"],
         "col": "Voertuigtype", "data": [("Totaal", "sum", "Som van Totaal")]},
    ]
    speed_cols = [h for h in out_header if h.startswith("Totaal snelheid ")]
    if speed_cols:
        data = [(c, "sum", f"Som {c}") for c in speed_cols]
        if "Gem." in out_header:
            data.append(("Gem.", "average", "Gemiddelde snelheid"))
        if "V85" in out_header:
            data.append(("V85", "average", "V85 (gem.)"))
        specs.append({"sheet": "Snelheidsverdeling per loc", "rows": ["Naam"],
                      "col": None, "data": data})

    geldig = []
    for s in specs:
        velden = list(s["rows"]) + ([s["col"]] if s["col"] else []) + [d[0] for d in s["data"]]
        if all(v in out_header for v in velden):
            geldig.append(s)
    return geldig


def _fmt_waarde(v):
    return v.strftime("%d-%m-%Y") if hasattr(v, "strftime") else v


def bereken_overzicht(out_header, out_rows, spec):
    """Bereken een statische kruistabel (terugval als Excel ontbreekt)."""
    idx = {h: i for i, h in enumerate(out_header)}
    rows_f, col_f, datas = spec["rows"], spec["col"], spec["data"]

    if col_f:
        veld = datas[0][0]
        col_values, rowkeys, cell = [], [], {}
        for row in out_rows:
            rk = tuple(_fmt_waarde(row[idx[f]]) for f in rows_f)
            cv = _fmt_waarde(row[idx[col_f]])
            if rk not in rowkeys:
                rowkeys.append(rk)
            if cv not in col_values:
                col_values.append(cv)
            x = row[idx[veld]]
            cell[(rk, cv)] = cell.get((rk, cv), 0) + (x if isinstance(x, (int, float)) else 0)
        header = list(rows_f) + [str(c) for c in col_values] + ["Eindtotaal"]
        table = [header]
        for rk in rowkeys:
            line = list(rk); som = 0
            for cv in col_values:
                v = cell.get((rk, cv), 0)
                line.append(v); som += v
            line.append(som)
            table.append(line)
        return table

    groups, rowkeys = {}, []
    for row in out_rows:
        rk = tuple(_fmt_waarde(row[idx[f]]) for f in rows_f)
        if rk not in rowkeys:
            rowkeys.append(rk)
        g = groups.setdefault(rk, {})
        for veld, _agg, _cap in datas:
            acc = g.setdefault(veld, [0, 0])
            x = row[idx[veld]]
            acc[0] += x if isinstance(x, (int, float)) else 0
            acc[1] += 1
    header = list(rows_f) + [cap for _, _, cap in datas]
    table = [header]
    for rk in rowkeys:
        line = list(rk)
        for veld, agg, _cap in datas:
            acc = groups[rk][veld]
            v = acc[0] if agg == "sum" else (acc[0] / acc[1] if acc[1] else 0)
            line.append(round(v, 2) if isinstance(v, float) else v)
        table.append(line)
    return table


def voeg_static_overzichten_toe(wb, out_header, out_rows):
    for spec in pivot_specs(out_header):
        ws = wb.create_sheet(title=spec["sheet"][:31])
        for r in bereken_overzicht(out_header, out_rows, spec):
            ws.append(r)


def voeg_pivots_toe_com(output_file, out_header, n_rows):
    """Laat Excel echte interactieve draaitabellen maken op extra tabbladen."""
    import gc
    xlDatabase, xlRowField, xlColumnField = 1, 1, 2
    xlSum, xlAverage = -4157, -4106
    specs = pivot_specs(out_header)
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = None
    try:
        wb = excel.Workbooks.Open(output_file)
        data_ws = wb.Worksheets("Samengevoegd")
        bron = data_ws.Range(data_ws.Cells(1, 1),
                             data_ws.Cells(n_rows + 1, len(out_header)))
        cache = wb.PivotCaches().Create(SourceType=xlDatabase, SourceData=bron)
        # In omgekeerde volgorde aanmaken: Add zet een nieuw blad vooraan,
        # zodat de bladen uiteindelijk in de juiste volgorde staan.
        for n, spec in enumerate(reversed(specs), start=1):
            ws = wb.Worksheets.Add()
            ws.Name = spec["sheet"]
            pt = cache.CreatePivotTable(TableDestination=ws.Cells(3, 1), TableName=f"Pivot{n}")
            for rf in spec["rows"]:
                pt.PivotFields(rf).Orientation = xlRowField
            if spec["col"]:
                pt.PivotFields(spec["col"]).Orientation = xlColumnField
            for veld, agg, caption in spec["data"]:
                pt.AddDataField(pt.PivotFields(veld), caption,
                                xlAverage if agg == "average" else xlSum)
        # Data-tabblad naar voren (positioneel; keyword Before/After bindt niet betrouwbaar)
        data_ws.Move(wb.Worksheets(1))
        wb.Save()
    finally:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        excel.Quit()
        excel = None
        gc.collect()


def schrijf_uitvoer(output_file, out_header, out_rows):
    """Schrijf de data weg en geef terug welke overzichten zijn toegevoegd."""
    datum_kolom = out_header.index("Datum") + 1
    if output_file.lower().endswith(".xlsx"):
        wb = Workbook(); ws = wb.active; ws.title = "Samengevoegd"
        ws.append(out_header)
        for row in out_rows:
            ws.append(row)
        for ri in range(2, len(out_rows) + 2):
            ws.cell(row=ri, column=datum_kolom).number_format = "DD-MM-YYYY"
        kleur_kolomtitels(ws, out_header)
        wb.save(output_file)
        # Draaitabellen (interactief via COM / statisch) zijn voorlopig uitgezet;
        # de functies hieronder blijven beschikbaar om later weer in te schakelen.
        return "geen"
    else:
        with open(output_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(out_header)
            for row in out_rows:
                csv_row = list(row)
                csv_row[datum_kolom - 1] = row[datum_kolom - 1].strftime("%d-%m-%Y")
                writer.writerow(csv_row)
        return "csv"


# ---------------------------------------------------------------------------
# Gatencontrole
# ---------------------------------------------------------------------------

def vind_gaten(all_rows, start_uur=6, eind_uur=24, min_run=2):
    """Zoek per (Locatie, Voertuigtype, Richting, Datum) naar runs van
    >= min_run aaneengesloten uren met 0 tellingen, binnen [start_uur, eind_uur)."""
    groepen = {}
    for d in all_rows:
        m = re.match(r"^(\d{1,2}):", str(d.get("Tijd", "")))
        if not m:
            continue
        uur = int(m.group(1))
        if uur < start_uur or uur >= eind_uur:
            continue
        key = (d.get("Code", ""), d.get("Naam", ""), d["Voertuigtype"], d["Richting"], d["Datum"])
        groepen.setdefault(key, {})
        groepen[key][uur] = groepen[key].get(uur, 0) + _num(d.get("Totaal", ""))

    gaten = []
    for key in sorted(groepen, key=lambda k: (k[1], k[2], k[4], k[3])):
        uren = groepen[key]
        run_start = None
        for uur in range(start_uur, eind_uur + 1):
            is_nul = uur < eind_uur and uren.get(uur, 0) == 0
            if is_nul and run_start is None:
                run_start = uur
            elif not is_nul and run_start is not None:
                if uur - run_start >= min_run:
                    gaten.append({"Code": key[0], "Naam": key[1], "Voertuigtype": key[2],
                                  "Richting": key[3], "Datum": key[4],
                                  "van": run_start, "tot": uur, "uren": uur - run_start})
                run_start = None
    return gaten


def gaten_overzicht_tekst(gaten):
    if not gaten:
        return "Geen gaten gevonden (06:00-24:00, minimaal 2 aaneengesloten lege uren)."
    lines = [f"{len(gaten)} mogelijke gaten gevonden",
             "(minimaal 2 aaneengesloten lege uren tussen 06:00 en 24:00):", ""]
    for g in gaten:
        datum = g["Datum"].strftime("%d-%m-%Y") if hasattr(g["Datum"], "strftime") else str(g["Datum"])
        loc = f"{g['Code']} {g['Naam']}".strip()
        lines.append(f"- {loc} | {g['Voertuigtype']} | {g['Richting']} | "
                     f"{datum} | {g['van']:02d}:00-{g['tot']:02d}:00 ({g['uren']} uur leeg)")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Grafische applicatie
# ---------------------------------------------------------------------------

class App:
    def __init__(self, root):
        self.root = root
        root.title("Verkeerstellingen samenvoegen")
        root.geometry("1180x700")
        root.minsize(980, 600)

        self.coords = {}   # path -> {naam, plaats, lat, lon}
        self.dd_rows = []

        root.rowconfigure(0, weight=1)
        root.columnconfigure(1, weight=1)

        links = tk.Frame(root)
        links.grid(row=0, column=0, sticky="ns", padx=8, pady=8)

        self._bouw_dagdelen(links)
        self.bu_zone = self._bouw_dropzone(
            links, "Gemotoriseerd vervoer",
            "dit zijn meestal de BU-bestanden", TELSLANG_VOERTUIGTYPE)
        self.fiets_zone = self._bouw_dropzone(
            links, "Fietsen",
            "dit zijn meestal de TU-xxx-fiets bestanden", FIETS_VOERTUIGTYPE)
        # aliassen zodat de overige code blijft werken
        self.bu_files = self.bu_zone["files"]
        self.fiets_files = self.fiets_zone["files"]
        # verplaats-knoppen naar het andere veld koppelen
        self.bu_zone["move_btn"].config(
            text="Verplaats selectie naar Fietsen",
            command=lambda: self._verplaats(self.bu_zone, self.fiets_zone))
        self.fiets_zone["move_btn"].config(
            text="Verplaats selectie naar Gemotoriseerd",
            command=lambda: self._verplaats(self.fiets_zone, self.bu_zone))

        # Kaart
        kaart_frame = tk.LabelFrame(root, text="Locaties (controle op compleetheid)")
        kaart_frame.grid(row=0, column=1, sticky="nsew", padx=8, pady=8)
        kaart_frame.rowconfigure(0, weight=1)
        kaart_frame.columnconfigure(0, weight=1)
        if HAS_MAP:
            self.map_widget = TkinterMapView(kaart_frame, corner_radius=0)
            self.map_widget.grid(row=0, column=0, sticky="nsew")
            self.map_widget.set_position(52.2, 5.3)  # midden NL
            self.map_widget.set_zoom(7)
        else:
            self.map_widget = None
            tk.Label(kaart_frame, text="(kaart niet beschikbaar)").grid(row=0, column=0)

        # Onderbalk
        onder = tk.Frame(root)
        onder.grid(row=1, column=0, columnspan=2, sticky="ew", padx=8, pady=(0, 8))
        self.status = tk.Label(onder, text="Sleep bestanden in de velden of gebruik 'Bladeren'.", anchor="w")
        self.status.pack(side="left", fill="x", expand=True)
        tk.Button(onder, text="Verwerken en opslaan", width=22,
                  command=self.verwerken).pack(side="right")
        tk.Button(onder, text="Controleer op gaten", width=20,
                  command=self.controleer_gaten).pack(side="right", padx=6)
        self.progress = ttk.Progressbar(onder, length=180, mode="determinate")
        self.progress.pack(side="right", padx=6)

    # ----- dagdelen -----
    def _bouw_dagdelen(self, parent):
        frame = tk.LabelFrame(parent, text="Dagdelen")
        frame.pack(fill="x", pady=(0, 8))
        kop = tk.Frame(frame); kop.pack(fill="x", padx=5, pady=(4, 0))
        tk.Label(kop, text="Van", width=6).grid(row=0, column=0)
        tk.Label(kop, text="Tot", width=6).grid(row=0, column=1)
        tk.Label(kop, text="Dagdeel", width=16).grid(row=0, column=2)
        self.dd_container = tk.Frame(frame); self.dd_container.pack(fill="x", padx=5)
        knoppen = tk.Frame(frame); knoppen.pack(fill="x", padx=5, pady=4)
        tk.Button(knoppen, text="+", width=2, command=lambda: self._dd_add()).pack(side="left")
        tk.Button(knoppen, text="Standaard", command=self._dd_standaard).pack(side="left", padx=4)
        self._dd_herbouw(lees_dagdelen_strings(basis_dir()))

    def _dd_add(self, van="", tot="", naam=""):
        rf = tk.Frame(self.dd_container); rf.pack(fill="x", pady=1)
        e_van = tk.Entry(rf, width=6); e_van.insert(0, van); e_van.grid(row=0, column=0, padx=1)
        e_tot = tk.Entry(rf, width=6); e_tot.insert(0, tot); e_tot.grid(row=0, column=1, padx=1)
        e_naam = tk.Entry(rf, width=16); e_naam.insert(0, naam); e_naam.grid(row=0, column=2, padx=1)
        entry = (e_van, e_tot, e_naam)

        def remove():
            if entry in self.dd_rows:
                self.dd_rows.remove(entry)
            rf.destroy()
        tk.Button(rf, text="x", width=2, command=remove).grid(row=0, column=3, padx=2)
        self.dd_rows.append(entry)

    def _dd_herbouw(self, rijen):
        for child in self.dd_container.winfo_children():
            child.destroy()
        self.dd_rows.clear()
        for van, tot, naam in rijen:
            self._dd_add(van, tot, naam)

    def _dd_standaard(self):
        self._dd_herbouw([tuple(r) for r in DEFAULT_DAGDELEN])

    def _dd_lees(self):
        rijen = []
        for e_van, e_tot, e_naam in self.dd_rows:
            van, tot, naam = e_van.get().strip(), e_tot.get().strip(), e_naam.get().strip()
            if not (van or tot or naam):
                continue
            if _tijd_naar_minuten(van) is None or _tijd_naar_minuten(tot) is None or not naam:
                messagebox.showerror("Ongeldige dagdeel-invoer",
                                     f"Controleer: van='{van}' tot='{tot}' dagdeel='{naam}'")
                return None
            rijen.append((van, tot, naam))
        if not rijen:
            messagebox.showerror("Dagdelen leeg", "Voeg minstens één dagdeel toe.")
            return None
        return rijen

    # ----- dropzones -----
    def _bouw_dropzone(self, parent, titel, hint, categorie):
        zone = {"files": [], "categorie": categorie, "naam": titel}
        frame = tk.LabelFrame(parent, text=titel)
        frame.pack(fill="both", expand=True, pady=(0, 8))
        tk.Label(frame, text=hint, fg="#555").pack(anchor="w", padx=6, pady=(2, 0))

        listbox = tk.Listbox(frame, width=40, height=6, selectmode="extended")
        listbox.pack(fill="both", expand=True, padx=6, pady=4)
        zone["listbox"] = listbox

        if HAS_DND:
            listbox.drop_target_register(DND_FILES)
            listbox.dnd_bind("<<Drop>>", lambda e, z=zone: self._op_drop(e, z))

        btns = tk.Frame(frame); btns.pack(fill="x", padx=6, pady=(0, 2))
        tk.Button(btns, text="Bladeren...",
                  command=lambda z=zone: self._bladeren(z)).pack(side="left")
        tk.Button(btns, text="Wissen",
                  command=lambda z=zone: self._wissen(z)).pack(side="left", padx=4)

        move_btn = tk.Button(frame, text="Verplaats selectie")
        move_btn.pack(fill="x", padx=6, pady=(0, 4))
        zone["move_btn"] = move_btn
        return zone

    def _op_drop(self, event, zone):
        self._voeg_toe(self.root.tk.splitlist(event.data), zone)

    def _bladeren(self, zone):
        paden = filedialog.askopenfilenames(
            title="Selecteer bestanden",
            filetypes=[("Excel/CSV", "*.xlsx *.csv"), ("Alle bestanden", "*.*")])
        self._voeg_toe(paden, zone)

    def _voeg_toe(self, paden, zone):
        for p in paden:
            p = str(p)
            if not (p.lower().endswith(".xlsx") or p.lower().endswith(".csv")):
                continue
            if p in zone["files"]:
                continue
            # typecontrole: lijkt het bestand op het andere type?
            self.status.config(text=f"Controleren: {Path(p).name}..."); self.root.update()
            gedetecteerd = detecteer_type(p)
            if gedetecteerd and gedetecteerd != zone["categorie"]:
                anders = ("een fietsbestand" if gedetecteerd == FIETS_VOERTUIGTYPE
                          else "een bestand voor gemotoriseerd vervoer")
                dit_veld = ("gemotoriseerd vervoer" if zone["categorie"] == TELSLANG_VOERTUIGTYPE
                            else "fietsen")
                if not messagebox.askyesno(
                        "Klopt het type?",
                        f"'{Path(p).name}' lijkt op {anders}.\n\n"
                        f"Weet je zeker dat dit een bestand voor {dit_veld} is?"):
                    continue
            zone["files"].append(p)
            zone["listbox"].insert("end", Path(p).name)
            if p.lower().endswith(".xlsx"):
                info = lees_info_velden(p)
                if info and info.get("lat") is not None:
                    self.coords[p] = info
        self._ververs_kaart()
        self.status.config(
            text=f"{len(self.bu_files)} gemotoriseerd, {len(self.fiets_files)} fiets — "
                 f"{len(self.coords)} locatie(s) op de kaart.")

    def _wissen(self, zone):
        zone["files"].clear()
        zone["listbox"].delete(0, "end")
        self._ververs_kaart()

    def _verplaats(self, bron, doel):
        sel = list(bron["listbox"].curselection())
        if not sel:
            messagebox.showinfo("Niets geselecteerd",
                                "Selecteer eerst één of meer bestanden in de lijst om te verplaatsen.")
            return
        verplaatst = 0
        for i in reversed(sel):
            path = bron["files"][i]
            del bron["files"][i]
            bron["listbox"].delete(i)
            if path not in doel["files"]:
                doel["files"].append(path)
                doel["listbox"].insert("end", Path(path).name)
                verplaatst += 1
        self._ververs_kaart()
        self.status.config(text=f"{verplaatst} bestand(en) verplaatst naar {doel['naam']}.")

    def _ververs_kaart(self):
        if not self.map_widget:
            return
        self.map_widget.delete_all_marker()
        actief = set(self.bu_files) | set(self.fiets_files)
        punten = []
        for p, info in self.coords.items():
            if p not in actief:
                continue
            label = info.get("naam") or Path(p).stem
            if info.get("plaats"):
                label = f"{label} ({info['plaats']})"
            self.map_widget.set_marker(info["lat"], info["lon"], text=label)
            punten.append((info["lat"], info["lon"]))
        if len(punten) == 1:
            self.map_widget.set_position(punten[0][0], punten[0][1]); self.map_widget.set_zoom(13)
        elif len(punten) > 1:
            lats = [a for a, _ in punten]; lons = [b for _, b in punten]
            self.map_widget.fit_bounding_box((max(lats), min(lons)), (min(lats), max(lons)))

    # ----- verwerken -----
    def _parse_met_progress(self):
        total = len(self.bu_files) + len(self.fiets_files)
        self.progress.config(maximum=max(total, 1), value=0)
        all_rows, telslang_headers, problemen = [], None, []
        idx = 0
        for path, label in ([(p, TELSLANG_VOERTUIGTYPE) for p in self.bu_files] +
                            [(p, FIETS_VOERTUIGTYPE) for p in self.fiets_files]):
            idx += 1
            self.status.config(text=f"Bezig met inlezen: bestand {idx} van {total} — {Path(path).name}")
            self.progress.config(value=idx)
            self.root.update()
            typed, headers = verwerk_bestand(path, label)
            if label == TELSLANG_VOERTUIGTYPE and headers and telslang_headers is None:
                telslang_headers = headers
            if not typed:
                soort = "gemotoriseerde" if label == TELSLANG_VOERTUIGTYPE else "fiets"
                problemen.append(f"- {Path(path).name} (geen {soort} data gevonden)")
            all_rows.extend(typed)
        return all_rows, telslang_headers, problemen

    def controleer_gaten(self):
        if not self.bu_files and not self.fiets_files:
            messagebox.showerror("Geen bestanden", "Voeg eerst bestanden toe.")
            return
        all_rows, _, _ = self._parse_met_progress()
        self.progress.config(value=0)
        if not all_rows:
            self.status.config(text="Geen data gevonden.")
            messagebox.showerror("Geen data", "Geen herkenbare data gevonden.")
            return
        gaten = vind_gaten(all_rows)
        self.status.config(text=f"Gatencontrole klaar: {len(gaten)} mogelijke gaten.")
        self._toon_gaten(gaten_overzicht_tekst(gaten), met_doorgaan=False)

    def verwerken(self):
        dd_strings = self._dd_lees()
        if dd_strings is None:
            return
        if not self.bu_files and not self.fiets_files:
            messagebox.showerror("Geen bestanden", "Voeg eerst bestanden toe.")
            return
        schrijf_dagdelen_strings(basis_dir(), dd_strings)
        dagdelen = strings_naar_dagdelen(dd_strings)

        all_rows, telslang_headers, problemen = self._parse_met_progress()
        self.progress.config(value=0)
        if not all_rows:
            self.status.config(text="Geen data gevonden.")
            messagebox.showerror("Geen data", "In de gekozen bestanden is geen herkenbare data gevonden.")
            return

        # Gatencontrole vóór het opslaan
        gaten = vind_gaten(all_rows)
        if gaten:
            self.status.config(text=f"{len(gaten)} mogelijke gaten gevonden — controleer.")
            doorgaan = self._toon_gaten(gaten_overzicht_tekst(gaten), met_doorgaan=True)
            if not doorgaan:
                self.status.config(text="Gestopt na gatencontrole.")
                return

        out_header, out_rows = build_output(all_rows, telslang_headers, dagdelen)
        n_fiets = sum(1 for d in all_rows if d["Voertuigtype"] == FIETS_VOERTUIGTYPE)
        n_tel = len(all_rows) - n_fiets

        output_file = filedialog.asksaveasfilename(
            title="Opslaan als", initialfile="samengevoegd.xlsx", defaultextension=".xlsx",
            filetypes=[("Excel bestanden", "*.xlsx"), ("CSV bestanden", "*.csv")])
        if not output_file:
            self.status.config(text="Opslaan geannuleerd.")
            return

        self.status.config(text="Bezig met opslaan..."); self.root.update()
        try:
            note = schrijf_uitvoer(output_file, out_header, out_rows)
        except PermissionError:
            self.status.config(text="Opslaan mislukt: bestand in gebruik.")
            messagebox.showerror("Bestand in gebruik",
                                 "Het doelbestand staat open (bv. in Excel). Sluit het en probeer opnieuw.")
            return

        self.status.config(text=f"Klaar: {len(out_rows)} rijen opgeslagen.")
        bericht = (f"{len(out_rows)} rijen geschreven naar:\n{output_file}\n\n"
                   f"({n_fiets} fiets, {n_tel} gemotoriseerd)\n")
        bericht += (f"\nGatencontrole: {len(gaten)} mogelijke gaten gevonden."
                    if gaten else "\nGatencontrole: geen gaten gevonden.")
        if note == "interactief":
            bericht += "\nDraaitabellen: interactief toegevoegd op extra tabbladen."
        elif note == "statisch":
            bericht += "\nDraaitabellen: statische overzichten toegevoegd (geen Excel-automatisering)."
        if problemen:
            bericht += "\n\nLet op, overgeslagen:\n" + "\n".join(problemen)
        messagebox.showinfo("Klaar", bericht)

    def _toon_gaten(self, tekst, met_doorgaan):
        dlg = tk.Toplevel(self.root)
        dlg.title("Gatencontrole (06:00 - 24:00)")
        dlg.geometry("720x470")
        dlg.grab_set()

        frame = tk.Frame(dlg)
        frame.pack(fill="both", expand=True, padx=8, pady=8)
        txt = tk.Text(frame, wrap="none")
        ysb = tk.Scrollbar(frame, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=ysb.set)
        ysb.pack(side="right", fill="y")
        txt.pack(side="left", fill="both", expand=True)
        txt.insert("1.0", tekst)
        txt.config(state="disabled")

        result = {"door": False}
        knoppen = tk.Frame(dlg)
        knoppen.pack(fill="x", padx=8, pady=(0, 8))

        def opslaan_txt():
            pad = filedialog.asksaveasfilename(
                title="Overzicht opslaan", defaultextension=".txt",
                initialfile="gatencontrole.txt", filetypes=[("Tekstbestand", "*.txt")])
            if pad:
                try:
                    with open(pad, "w", encoding="utf-8") as f:
                        f.write(tekst)
                except OSError:
                    pass

        tk.Button(knoppen, text="Opslaan naar bestand", command=opslaan_txt).pack(side="left")
        if met_doorgaan:
            def door():
                result["door"] = True
                dlg.destroy()
            tk.Button(knoppen, text="Annuleren", command=dlg.destroy).pack(side="right")
            tk.Button(knoppen, text="Doorgaan met opslaan", command=door).pack(side="right", padx=6)
        else:
            tk.Button(knoppen, text="Sluiten", command=dlg.destroy).pack(side="right")

        dlg.wait_window()
        return result["door"]


def main():
    root = TkinterDnD.Tk() if HAS_DND else tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        import traceback
        traceback.print_exc()
        try:
            messagebox.showerror("Fout", f"Er ging iets mis:\n\n{exc}")
        except Exception:
            pass
