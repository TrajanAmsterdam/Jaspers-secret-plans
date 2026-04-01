import csv
import re
import tkinter as tk
from tkinter import filedialog, Listbox, Toplevel, MULTIPLE, END
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook


def read_csv_rows(file_path):
    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter=";")
        return list(reader)


def read_xlsx_rows(file_path, sheet_name):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([cell if cell is not None else "" for cell in row])
    wb.close()
    return rows


def select_sheets(file_path):
    """Open een dialoog om tabbladen te kiezen uit een xlsx bestand."""
    wb = load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    if len(sheet_names) == 1:
        return sheet_names

    dialog = Toplevel()
    dialog.title(f"Tabbladen - {Path(file_path).name}")
    dialog.geometry("450x400")
    dialog.minsize(350, 250)
    dialog.grab_set()

    tk.Label(dialog, text="Selecteer tabbladen om samen te voegen (Ctrl+klik voor meerdere):").pack(pady=5)
    listbox = Listbox(dialog, selectmode=MULTIPLE, width=50, height=15)
    listbox.pack(padx=10, pady=5, fill="both", expand=True)
    for name in sheet_names:
        listbox.insert(END, name)

    selected = []

    def on_ok():
        for idx in listbox.curselection():
            selected.append(sheet_names[idx])
        dialog.destroy()

    btn_frame = tk.Frame(dialog)
    btn_frame.pack(pady=10)
    tk.Button(btn_frame, text="OK", command=on_ok, width=15).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Annuleren", command=dialog.destroy, width=15).pack(side="left", padx=5)
    dialog.wait_window()
    return selected


def to_number(value):
    """Converteer een waarde naar een getal. Strings met komma worden decimaalteken."""
    if isinstance(value, (int, float)):
        return value
    if not isinstance(value, str):
        return value
    value = value.strip()
    if not value:
        return ""
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value.replace(",", "."))
    except ValueError:
        return value


def parse_rows(rows):
    """Parse de rijen van een enkele tabel (CSV of xlsx sheet) en geeft data-rijen terug."""
    # Build combined column headers from first table's header rows
    header1 = None
    header2 = None
    for i, row in enumerate(rows):
        if row and str(row[0]).strip() == "Tijd":
            header1 = [str(c) for c in row]
            header2 = [str(c) for c in rows[i + 1]]
            break

    if header1 is None:
        return None, []

    # Combine the two header rows for the speed/length columns (indices 1-24)
    combined_headers = []
    for j in range(1, 25):
        length_class = header1[j].strip()
        speed_class = header2[j].strip()
        combined_headers.append(f"{length_class};{speed_class}")

    # Parse each table block
    table_header_pattern = re.compile(r"^(\d{2}-\d{2}-\d{4})\s*-\s*(.+)$")
    data_rows = []

    i = 0
    while i < len(rows):
        row = rows[i]
        if row and row[0]:
            match = table_header_pattern.match(str(row[0]).strip())
            if match:
                datum = datetime.strptime(match.group(1), "%d-%m-%Y")
                richting = match.group(2).strip()
                i += 3  # table header + header1 + header2
                while i < len(rows) and rows[i] and str(rows[i][0]).strip():
                    data_row = rows[i]
                    tijd = str(data_row[0]).strip()
                    values = [to_number(data_row[j]) for j in range(1, 25)]
                    v85 = to_number(data_row[27]) if len(data_row) > 27 else ""
                    gem = to_number(data_row[29]) if len(data_row) > 29 else ""
                    totaal = to_number(data_row[30]) if len(data_row) > 30 else ""
                    data_rows.append([datum, tijd, richting] + values + [v85, gem, totaal])
                    i += 1
                continue
        i += 1

    return combined_headers, data_rows


def main():
    root = tk.Tk()
    root.withdraw()

    print("Selecteer invoerbestand(en)...")
    input_files = filedialog.askopenfilenames(
        title="Selecteer invoerbestanden",
        filetypes=[
            ("CSV en Excel bestanden", "*.csv *.xlsx"),
            ("CSV bestanden", "*.csv"),
            ("Excel bestanden", "*.xlsx"),
            ("Alle bestanden", "*.*"),
        ],
    )
    if not input_files:
        print("Geen bestanden geselecteerd, script gestopt.")
        return

    all_data_rows = []
    combined_headers = None

    for file_path in input_files:
        path = Path(file_path)
        locatie = path.stem
        print(f"Verwerken: {path.name}")

        if path.suffix.lower() == ".xlsx":
            sheets = select_sheets(file_path)
            if not sheets:
                print(f"  Geen tabbladen geselecteerd voor {path.name}, overgeslagen.")
                continue
            for sheet in sheets:
                print(f"  Tabblad: {sheet}")
                rows = read_xlsx_rows(file_path, sheet)
                headers, data = parse_rows(rows)
                if headers and combined_headers is None:
                    combined_headers = headers
                for row in data:
                    all_data_rows.append([locatie] + row)
        else:
            rows = read_csv_rows(file_path)
            headers, data = parse_rows(rows)
            if headers and combined_headers is None:
                combined_headers = headers
            for row in data:
                all_data_rows.append([locatie] + row)

    if not all_data_rows:
        print("Geen data gevonden in de geselecteerde bestanden.")
        return

    out_header = ["Locatie", "Datum", "Tijd", "Richting"] + combined_headers + ["V85", "Gem.", "Totaal"]

    print("Selecteer opslaglocatie voor het uitvoerbestand...")
    output_file = filedialog.asksaveasfilename(
        title="Opslaan als",
        initialdir=str(Path(input_files[0]).parent),
        initialfile="samengevoegd.xlsx",
        defaultextension=".xlsx",
        filetypes=[("Excel bestanden", "*.xlsx"), ("CSV bestanden", "*.csv"), ("Alle bestanden", "*.*")],
    )
    if not output_file:
        print("Geen opslaglocatie geselecteerd, script gestopt.")
        return

    if output_file.lower().endswith(".xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Samengevoegd"
        ws.append(out_header)
        for row in all_data_rows:
            ws.append(row)
        # Formatteer datumkolom (kolom B) als datum
        for row_idx in range(2, len(all_data_rows) + 2):
            ws.cell(row=row_idx, column=2).number_format = "DD-MM-YYYY"
        wb.save(output_file)
    else:
        with open(output_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(out_header)
            for row in all_data_rows:
                csv_row = [row[0], row[1].strftime("%d-%m-%Y")] + row[2:]
                writer.writerow(csv_row)

    print(f"Klaar! {len(all_data_rows)} rijen geschreven naar {output_file}")


if __name__ == "__main__":
    main()
