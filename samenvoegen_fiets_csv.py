import csv
import re
import tkinter as tk
from tkinter import filedialog, Listbox, Toplevel, MULTIPLE, END
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


# Mapping van kolomnaam naar weekdag-offset (maandag=0)
DAG_KOLOMMEN = {
    "Maandag": 0,
    "Dinsdag": 1,
    "Woensdag": 2,
    "Donderdag": 3,
    "Vrijdag": 4,
    "Zaterdag": 5,
    "Zondag": 6,
}


def read_csv_rows(file_path):
    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter=";")
        return list(reader)


def read_xlsx_rows(file_path, sheet_name):
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([cell if cell is not None else "" for cell in row])
    wb.close()
    return rows


def select_sheets(file_path):
    """Open een dialoog om tabbladen te kiezen uit een xlsx bestand."""
    wb = load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    if len(sheet_names) == 1:
        return sheet_names

    dialog = Toplevel()
    dialog.title(f"Tabbladen - {Path(file_path).name}")
    dialog.geometry("450x400")
    dialog.minsize(350, 250)
    dialog.grab_set()

    tk.Label(dialog, text="Selecteer tabbladen om samen te voegen (Ctrl+klik voor meerdere):").pack(pady=5)
    listbox = Listbox(dialog, selectmode=MULTIPLE, width=50, height=15)
    listbox.pack(padx=10, pady=5, fill="both", expand=True)
    for name in sheet_names:
        listbox.insert(END, name)

    selected = []

    def on_ok():
        for idx in listbox.curselection():
            selected.append(sheet_names[idx])
        dialog.destroy()

    btn_frame = tk.Frame(dialog)
    btn_frame.pack(pady=10)
    tk.Button(btn_frame, text="OK", command=on_ok, width=15).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Annuleren", command=dialog.destroy, width=15).pack(side="left", padx=5)
    dialog.wait_window()
    return selected


def to_number(value):
    """Converteer een waarde naar een getal."""
    if isinstance(value, (int, float)):
        return value
    if not isinstance(value, str):
        return value
    value = value.strip()
    if not value:
        return 0
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value.replace(",", "."))
    except ValueError:
        return value


def parse_week_start(header_text):
    """Parse de startdatum uit een tabelheader zoals '16-02-2026 - 22-02-2026 - Hellingweg richting Purmerhoek'."""
    match = re.match(r"^(\d{2}-\d{2}-\d{4})\s*-\s*\d{2}-\d{2}-\d{4}\s*-\s*(.+)$", header_text.strip())
    if match:
        start_date = datetime.strptime(match.group(1), "%d-%m-%Y")
        richting = match.group(2).strip()
        return start_date, richting
    return None, None


def parse_rows(rows):
    """Parse de rijen en klap weektabellen uit naar individuele dagen."""
    # Zoek de header-rij om kolomposities te bepalen
    header_row = None
    for i, row in enumerate(rows):
        if row and str(row[0]).strip() == "Tijd":
            header_row = row
            break

    if header_row is None:
        return []

    # Bepaal kolom-indices voor elke dag (alleen absolute aantallen, niet %)
    dag_indices = {}
    for j, col in enumerate(header_row):
        col_name = str(col).strip()
        if col_name in DAG_KOLOMMEN:
            dag_indices[col_name] = j

    # Parse tabelblokken
    table_header_pattern = re.compile(
        r"^(\d{2}-\d{2}-\d{4})\s*-\s*\d{2}-\d{2}-\d{4}\s*-\s*(.+)$"
    )
    data_rows = []

    i = 0
    while i < len(rows):
        row = rows[i]
        if row and row[0]:
            match = table_header_pattern.match(str(row[0]).strip())
            if match:
                start_date = datetime.strptime(match.group(1), "%d-%m-%Y")
                richting = match.group(2).strip()
                i += 2  # skip header + kolom-headerrij
                # Lees alle uurregels van dit blok
                uur_rijen = []
                while i < len(rows) and rows[i] and str(rows[i][0]).strip():
                    uur_rijen.append(rows[i])
                    i += 1
                # Uitklappen per dag, dan per uur
                for dag_naam, offset in DAG_KOLOMMEN.items():
                    if dag_naam in dag_indices:
                        datum = start_date + timedelta(days=offset)
                        for data_row in uur_rijen:
                            tijd = str(data_row[0]).strip()
                            aantal = to_number(data_row[dag_indices[dag_naam]])
                            data_rows.append([datum, tijd, richting, aantal])
                continue
        i += 1

    return data_rows


def main():
    root = tk.Tk()
    root.withdraw()

    print("Selecteer invoerbestand(en)...")
    input_files = filedialog.askopenfilenames(
        title="Selecteer invoerbestanden",
        filetypes=[
            ("CSV en Excel bestanden", "*.csv *.xlsx"),
            ("CSV bestanden", "*.csv"),
            ("Excel bestanden", "*.xlsx"),
            ("Alle bestanden", "*.*"),
        ],
    )
    if not input_files:
        print("Geen bestanden geselecteerd, script gestopt.")
        return

    all_data_rows = []

    for file_path in input_files:
        path = Path(file_path)
        locatie = path.stem
        print(f"Verwerken: {path.name}")

        if path.suffix.lower() == ".xlsx":
            sheets = select_sheets(file_path)
            if not sheets:
                print(f"  Geen tabbladen geselecteerd voor {path.name}, overgeslagen.")
                continue
            for sheet in sheets:
                print(f"  Tabblad: {sheet}")
                rows = read_xlsx_rows(file_path, sheet)
                data = parse_rows(rows)
                for row in data:
                    all_data_rows.append([locatie] + row)
        else:
            rows = read_csv_rows(file_path)
            data = parse_rows(rows)
            for row in data:
                all_data_rows.append([locatie] + row)

    if not all_data_rows:
        print("Geen data gevonden in de geselecteerde bestanden.")
        return

    out_header = ["Locatie", "Datum", "Tijd", "Richting", "Aantal"]

    print(f"{len(all_data_rows)} rijen verwerkt.")
    print("Selecteer opslaglocatie voor het uitvoerbestand...")
    output_file = filedialog.asksaveasfilename(
        title="Opslaan als",
        initialdir=str(Path(input_files[0]).parent),
        initialfile="samengevoegd_fiets.xlsx",
        defaultextension=".xlsx",
        filetypes=[("Excel bestanden", "*.xlsx"), ("CSV bestanden", "*.csv"), ("Alle bestanden", "*.*")],
    )
    if not output_file:
        print("Geen opslaglocatie geselecteerd, script gestopt.")
        return

    if output_file.lower().endswith(".xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Samengevoegd"
        ws.append(out_header)
        for row in all_data_rows:
            ws.append(row)
        # Formatteer datumkolom (kolom B) als datum
        for row_idx in range(2, len(all_data_rows) + 2):
            ws.cell(row=row_idx, column=2).number_format = "DD-MM-YYYY"
        wb.save(output_file)
    else:
        with open(output_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(out_header)
            # Voor CSV: datums terug naar string
            for row in all_data_rows:
                csv_row = [row[0], row[1].strftime("%d-%m-%Y")] + row[2:]
                writer.writerow(csv_row)

    print(f"Klaar! {len(all_data_rows)} rijen geschreven naar {output_file}")


if __name__ == "__main__":
    main()
